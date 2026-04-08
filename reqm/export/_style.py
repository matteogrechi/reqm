"""Shared Excel formatting helpers for all exporters."""
from __future__ import annotations
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, PatternFill, Alignment

HEADER_FILL = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
BOLD_FONT = Font(bold=True)


def apply_header_style(ws: Worksheet, tab_colour: str) -> None:
    """Apply bold font, #BDD7EE fill, freeze pane A2, and tab colour to worksheet.

    Args:
        ws: The openpyxl worksheet to style.
        tab_colour: Colour name or hex string for the sheet tab.
    """
    ws.sheet_properties.tabColor = tab_colour
    ws.freeze_panes = "A2"

    for cell in ws[1]:
        cell.font = BOLD_FONT
        cell.fill = HEADER_FILL
