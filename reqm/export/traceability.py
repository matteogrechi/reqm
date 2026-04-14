"""Built-in exporter: traceability matrix as ECSS-styled Excel."""
from __future__ import annotations
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment

from reqm.export.base import AbstractExporter
from reqm.export._style import apply_header_style
from reqm.models import Requirement, FolderMeta


class TraceabilityExporter(AbstractExporter):
    name = "traceability"
    description = "ECSS traceability matrix (parent → children links)"

    def export(
        self,
        requirements: list[Requirement],
        folders: list[FolderMeta],
        output: Path,
    ) -> None:
        """Write an ECSS traceability matrix workbook.

        Sheets produced:

        - **Matrix**: rows are source requirements; columns are deriving
          requirement IDs; cells contain ✓ where a derived_from link exists.
        - **Orphans**: requirements with no derived_from and no related_to.

        Args:
            requirements: Full collection of validated requirements.
            folders: Folder metadata for all discovered folders.
            output: Destination path for the .xlsx file.
        """
        wb = Workbook()

        # --- Matrix sheet ---
        ws_matrix = wb.active
        ws_matrix.title = "Matrix"

        # Column headers: first cell blank, then every requirement id
        child_ids = [r.id for r in requirements]
        header_row = [""] + child_ids
        ws_matrix.append(header_row)

        # Build lookup: id → requirement
        req_by_id = {r.id: r for r in requirements}

        for req in requirements:
            row = [req.id]
            for child_id in child_ids:
                child = req_by_id.get(child_id)
                if child and req.id in child.derived_from:
                    row.append("✓")
                else:
                    row.append("")
            ws_matrix.append(row)

        apply_header_style(ws_matrix, "4472C4")

        # Set column widths
        ws_matrix.column_dimensions["A"].width = 16
        for i in range(2, len(child_ids) + 2):
            col_letter = chr(64 + i)  # A=1, B=2, ...
            ws_matrix.column_dimensions[col_letter].width = 14
            # Center-align cells in data columns
            for row_idx in range(2, len(requirements) + 2):
                ws_matrix.cell(row=row_idx, column=i).alignment = Alignment(horizontal="center")

        # --- Orphans sheet ---
        ws_orphans = wb.create_sheet("Orphans")
        ws_orphans.append(["ID", "Title", "Type"])

        for req in requirements:
            if not req.derived_from and not req.related_to:
                ws_orphans.append([req.id, req.title, req.type])

        apply_header_style(ws_orphans, "BF8F00")
        ws_orphans.column_dimensions["A"].width = 16
        ws_orphans.column_dimensions["B"].width = 35
        ws_orphans.column_dimensions["C"].width = 16

        wb.save(str(output))
