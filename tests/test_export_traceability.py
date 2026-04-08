"""Tests for reqm.export.traceability — matrix values, orphans."""
from pathlib import Path

from openpyxl import load_workbook

from reqm.export.traceability import TraceabilityExporter
from reqm.fs import load_requirements, load_folder_meta


def test_traceability_exporter(fixtures_dir: Path, tmp_path: Path):
    reqs = load_requirements(fixtures_dir)
    folders = []
    meta = load_folder_meta(fixtures_dir)
    if meta:
        folders.append(meta)

    output = tmp_path / "traceability.xlsx"
    exporter = TraceabilityExporter()
    exporter.export(reqs, folders, output)

    assert output.exists()

    wb = load_workbook(str(output))
    assert "Matrix" in wb.sheetnames
    assert "Orphans" in wb.sheetnames


def test_matrix_has_checkmarks(fixtures_dir: Path, tmp_path: Path):
    reqs = load_requirements(fixtures_dir)
    folders = []
    meta = load_folder_meta(fixtures_dir)
    if meta:
        folders.append(meta)

    output = tmp_path / "trace.xlsx"
    exporter = TraceabilityExporter()
    exporter.export(reqs, folders, output)

    wb = load_workbook(str(output))
    ws = wb["Matrix"]

    # Rows are parents, columns are children.
    # B derived_from A → cell at row for A, column for B should have ✓
    child_ids = [cell.value for cell in ws[1][1:]]  # skip first empty cell
    row_map = {}
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=False):
        row_map[row[0].value] = row[0].row

    # B's derived_from is A, so row A, column B should have ✓
    if "A" in row_map and "B" in child_ids:
        col_b = child_ids.index("B") + 2
        cell = ws.cell(row=row_map["A"], column=col_b)
        assert cell.value == "✓"


def test_orphans_sheet(fixtures_dir: Path, tmp_path: Path):
    reqs = load_requirements(fixtures_dir)
    folders = []
    meta = load_folder_meta(fixtures_dir)
    if meta:
        folders.append(meta)

    output = tmp_path / "orphans.xlsx"
    exporter = TraceabilityExporter()
    exporter.export(reqs, folders, output)

    wb = load_workbook(str(output))
    ws = wb["Orphans"]

    orphan_ids = set()
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=False):
        orphan_ids.add(row[0].value)

    # D is an orphan (no derived_from, no related_to)
    assert "D" in orphan_ids
    # B is not an orphan (has derived_from and related_to)
    assert "B" not in orphan_ids


def test_traceability_formatting(fixtures_dir: Path, tmp_path: Path):
    reqs = load_requirements(fixtures_dir)
    folders = []
    meta = load_folder_meta(fixtures_dir)
    if meta:
        folders.append(meta)

    output = tmp_path / "trace_fmt.xlsx"
    exporter = TraceabilityExporter()
    exporter.export(reqs, folders, output)

    wb = load_workbook(str(output))
    ws_matrix = wb["Matrix"]
    ws_orphans = wb["Orphans"]

    # Header row is bold
    for cell in ws_matrix[1]:
        assert cell.font.bold is True

    # Tab colours set
    assert ws_matrix.sheet_properties.tabColor is not None
    assert ws_orphans.sheet_properties.tabColor is not None
