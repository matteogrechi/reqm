"""Tests for reqm.export.requirements — sheet names, columns, formatting."""
from pathlib import Path

from openpyxl import load_workbook

from reqm.export.requirements import RequirementsExporter
from reqm.fs import load_requirements, load_folder_meta


def test_requirements_exporter(fixtures_dir: Path, tmp_path: Path):
    reqs = load_requirements(fixtures_dir)
    folders = []
    meta = load_folder_meta(fixtures_dir)
    if meta:
        folders.append(meta)

    output = tmp_path / "requirements.xlsx"
    exporter = RequirementsExporter()
    exporter.export(reqs, folders, [], output)

    assert output.exists()

    wb = load_workbook(str(output))
    assert "Requirements" in wb.sheetnames


def test_requirements_sheet_headers(fixtures_dir: Path, tmp_path: Path):
    reqs = load_requirements(fixtures_dir)
    folders = []
    meta = load_folder_meta(fixtures_dir)
    if meta:
        folders.append(meta)

    output = tmp_path / "req_hdr.xlsx"
    exporter = RequirementsExporter()
    exporter.export(reqs, folders, [], output)

    wb = load_workbook(str(output))
    ws = wb["Requirements"]
    headers = [cell.value for cell in ws[1]]
    assert headers == [
        "Is Folder", "ID", "Parent", "Title", "Description", "Rationale",
        "Acceptance Criteria", "Type", "Verification Method",
        "Derived From", "Related To", "Tags", "Path",
    ]


def test_requirements_sheet_has_data_rows(fixtures_dir: Path, tmp_path: Path):
    reqs = load_requirements(fixtures_dir)
    folders = []
    meta = load_folder_meta(fixtures_dir)
    if meta:
        folders.append(meta)

    output = tmp_path / "req_data.xlsx"
    exporter = RequirementsExporter()
    exporter.export(reqs, folders, [], output)

    wb = load_workbook(str(output))
    ws = wb["Requirements"]
    assert ws.max_row > 1


def test_requirements_formatting(fixtures_dir: Path, tmp_path: Path):
    reqs = load_requirements(fixtures_dir)
    folders = []
    meta = load_folder_meta(fixtures_dir)
    if meta:
        folders.append(meta)

    output = tmp_path / "req_fmt.xlsx"
    exporter = RequirementsExporter()
    exporter.export(reqs, folders, [], output)

    wb = load_workbook(str(output))
    ws = wb["Requirements"]

    # Header row is bold
    for cell in ws[1]:
        assert cell.font.bold is True

    # Freeze pane set
    assert ws.freeze_panes == "A2"

    # Tab colour set
    assert ws.sheet_properties.tabColor is not None