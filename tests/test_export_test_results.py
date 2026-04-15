"""Tests for reqm.export.test_results — sheet names, item count, coverage."""
from pathlib import Path

from openpyxl import load_workbook

from reqm.export.test_results import TestResultsExporter
from reqm.fs import load_requirements, load_folder_meta, load_validation_items


def test_test_results_exporter(fixtures_dir: Path, validation_items_dir: Path, tmp_path: Path):
    reqs = load_requirements(fixtures_dir)
    folders = []
    meta = load_folder_meta(fixtures_dir)
    if meta:
        folders.append(meta)
    items = load_validation_items(validation_items_dir)

    output = tmp_path / "test_results.xlsx"
    exporter = TestResultsExporter()
    exporter.export(reqs, folders, items, output)

    assert output.exists()

    wb = load_workbook(str(output))
    assert "Results" in wb.sheetnames
    assert "Coverage" in wb.sheetnames


def test_results_sheet_headers(fixtures_dir: Path, validation_items_dir: Path, tmp_path: Path):
    reqs = load_requirements(fixtures_dir)
    folders = []
    meta = load_folder_meta(fixtures_dir)
    if meta:
        folders.append(meta)
    items = load_validation_items(validation_items_dir)

    output = tmp_path / "results_hdr.xlsx"
    exporter = TestResultsExporter()
    exporter.export(reqs, folders, items, output)

    wb = load_workbook(str(output))
    ws = wb["Results"]
    headers = [cell.value for cell in ws[1]]
    assert headers == ["Req ID", "Item ID", "Item Title"]


def test_coverage_sheet_headers(fixtures_dir: Path, validation_items_dir: Path, tmp_path: Path):
    reqs = load_requirements(fixtures_dir)
    folders = []
    meta = load_folder_meta(fixtures_dir)
    if meta:
        folders.append(meta)
    items = load_validation_items(validation_items_dir)

    output = tmp_path / "coverage_hdr.xlsx"
    exporter = TestResultsExporter()
    exporter.export(reqs, folders, items, output)

    wb = load_workbook(str(output))
    ws = wb["Coverage"]
    headers = [cell.value for cell in ws[1]]
    assert headers == ["Req ID", "Title", "Verified (Y/N)", "Verification Method", "Item Count"]


def test_coverage_calc_zero_items(fixtures_dir: Path, validation_items_dir: Path, tmp_path: Path):
    """Requirements with no validated_by entries show Item Count = 0."""
    reqs = load_requirements(fixtures_dir)
    folders = []
    meta = load_folder_meta(fixtures_dir)
    if meta:
        folders.append(meta)
    items = load_validation_items(validation_items_dir)

    output = tmp_path / "coverage_zero.xlsx"
    exporter = TestResultsExporter()
    exporter.export(reqs, folders, items, output)

    wb = load_workbook(str(output))
    ws = wb["Coverage"]

    item_count_map = {}
    for row in ws.iter_rows(min_row=2, values_only=False):
        req_id = row[0].value
        item_count_map[req_id] = row[4].value

    assert item_count_map.get("A") == 0
    assert item_count_map.get("D") == 0


def test_coverage_calc_with_items(fixtures_dir: Path, validation_items_dir: Path, tmp_path: Path):
    """Requirements with validated_by entries show correct Item Count and Verified flag."""
    reqs = load_requirements(fixtures_dir)
    folders = []
    meta = load_folder_meta(fixtures_dir)
    if meta:
        folders.append(meta)
    items = load_validation_items(validation_items_dir)

    output = tmp_path / "coverage_with_items.xlsx"
    exporter = TestResultsExporter()
    exporter.export(reqs, folders, items, output)

    wb = load_workbook(str(output))
    ws = wb["Coverage"]

    item_count_map = {}
    verified_map = {}
    for row in ws.iter_rows(min_row=2, values_only=False):
        req_id = row[0].value
        verified_map[req_id] = row[2].value
        item_count_map[req_id] = row[4].value

    assert verified_map.get("E") == "Y"
    assert verified_map.get("A") == "N"
    assert item_count_map.get("E") == 2
    assert item_count_map.get("A") == 0


def test_results_sheet_rows(fixtures_dir: Path, validation_items_dir: Path, tmp_path: Path):
    """Results sheet contains one row per linked item with resolved title."""
    reqs = load_requirements(fixtures_dir)
    folders = []
    meta = load_folder_meta(fixtures_dir)
    if meta:
        folders.append(meta)
    items = load_validation_items(validation_items_dir)

    output = tmp_path / "results_rows.xlsx"
    exporter = TestResultsExporter()
    exporter.export(reqs, folders, items, output)

    wb = load_workbook(str(output))
    ws = wb["Results"]

    rows = [(row[0].value, row[1].value, row[2].value)
            for row in ws.iter_rows(min_row=2, values_only=False)]

    # E links TC-001 and TC-002
    e_rows = [(rid, iid, title) for rid, iid, title in rows if rid == "E"]
    assert len(e_rows) == 2
    item_ids = {r[1] for r in e_rows}
    assert item_ids == {"TC-001", "TC-002"}
    # Titles should be resolved from the ValidationItem objects
    for _, iid, title in e_rows:
        assert title != ""


def test_test_results_formatting(fixtures_dir: Path, validation_items_dir: Path, tmp_path: Path):
    reqs = load_requirements(fixtures_dir)
    folders = []
    meta = load_folder_meta(fixtures_dir)
    if meta:
        folders.append(meta)
    items = load_validation_items(validation_items_dir)

    output = tmp_path / "results_fmt.xlsx"
    exporter = TestResultsExporter()
    exporter.export(reqs, folders, items, output)

    wb = load_workbook(str(output))
    ws_results = wb["Results"]
    ws_coverage = wb["Coverage"]

    for cell in ws_results[1]:
        assert cell.font.bold is True
    for cell in ws_coverage[1]:
        assert cell.font.bold is True

    assert ws_results.sheet_properties.tabColor is not None
    assert ws_coverage.sheet_properties.tabColor is not None