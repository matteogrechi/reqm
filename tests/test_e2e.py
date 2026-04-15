"""End-to-end tests for the reqm CLI.

Each test exercises the CLI from the outside via Click's CliRunner, asserting
on exit codes, stdout content, and (for exports) the resulting Excel workbook.
"""
from __future__ import annotations

import textwrap
from pathlib import Path

import openpyxl
import pytest
from click.testing import CliRunner

from reqm.cli import cli


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _req(content: str) -> str:
    """Strip common indentation from an inline requirement string."""
    return textwrap.dedent(content)


def _make_collection(root: Path, reqs: list[tuple[str, str]]) -> None:
    """Seed root with a .folder-metadata.md and one file per (name, content) pair."""
    (root / ".folder-metadata.md").write_text(
        "---\nid: TST\ntitle: Test Collection\n---\n"
    )
    for filename, content in reqs:
        (root / filename).write_text(_req(content))


def _data_lines(output: str) -> list[str]:
    """Return non-empty lines after the header and separator rows."""
    lines = output.splitlines()
    return [l for l in lines[2:] if l.strip()]


# ---------------------------------------------------------------------------
# list
# ---------------------------------------------------------------------------

def test_list_column_headers(fixtures_dir: Path) -> None:
    result = CliRunner().invoke(cli, ["list", "--root", str(fixtures_dir)])
    assert result.exit_code == 0
    for header in ("ID", "Title", "Type", "Verification", "Folder ID"):
        assert header in result.output


def test_list_shows_all_requirements(fixtures_dir: Path) -> None:
    result = CliRunner().invoke(cli, ["list", "--root", str(fixtures_dir)])
    assert result.exit_code == 0
    rows = _data_lines(result.output)
    ids = [r.split()[0] for r in rows]
    for req_id in ("A", "B", "C", "D", "E", "F"):
        assert req_id in ids, f"requirement {req_id!r} missing from list output"


def test_list_folder_filter_returns_only_matching(fixtures_dir: Path) -> None:
    result = CliRunner().invoke(cli, ["list", "--root", str(fixtures_dir), "--folder", "GA"])
    assert result.exit_code == 0
    rows = _data_lines(result.output)
    ids = [r.split()[0] for r in rows]
    assert ids == ["F"]


def test_list_status_filter_includes_match(tmp_path: Path) -> None:
    _make_collection(tmp_path, [
        ("R1-active.md", """\
            ---
            id: R1
            title: Active req
            type: Functional
            verification: [Test]
            status: Active
            ---
            """),
        ("R2-draft.md", """\
            ---
            id: R2
            title: Draft req
            type: Functional
            verification: [Test]
            status: Draft
            ---
            """),
    ])
    result = CliRunner().invoke(cli, ["list", "--root", str(tmp_path), "--status", "Active"])
    assert result.exit_code == 0
    ids = [r.split()[0] for r in _data_lines(result.output)]
    assert "R1" in ids
    assert "R2" not in ids


def test_list_status_filter_no_match_returns_empty(tmp_path: Path) -> None:
    _make_collection(tmp_path, [
        ("R1.md", """\
            ---
            id: R1
            title: A req
            type: Functional
            verification: [Test]
            ---
            """),
    ])
    result = CliRunner().invoke(cli, ["list", "--root", str(tmp_path), "--status", "Deprecated"])
    assert result.exit_code == 0
    assert _data_lines(result.output) == []


# ---------------------------------------------------------------------------
# show
# ---------------------------------------------------------------------------

def test_show_required_fields(fixtures_dir: Path) -> None:
    result = CliRunner().invoke(cli, ["show", "A", "--root", str(fixtures_dir)])
    assert result.exit_code == 0
    for label in ("ID:", "Title:", "Type:", "Verification:"):
        assert label in result.output


def test_show_body_sections(fixtures_dir: Path) -> None:
    result = CliRunner().invoke(cli, ["show", "A", "--root", str(fixtures_dir)])
    assert result.exit_code == 0
    for section in ("Description:", "Rationale:", "Acceptance Criteria:"):
        assert section in result.output


def test_show_displays_derived_from_and_related_to(fixtures_dir: Path) -> None:
    # B is derived from A and related to C
    result = CliRunner().invoke(cli, ["show", "B", "--root", str(fixtures_dir)])
    assert result.exit_code == 0
    assert "Derived From:" in result.output
    assert "A" in result.output
    assert "Related To:" in result.output
    assert "C" in result.output


def test_show_displays_tags(fixtures_dir: Path) -> None:
    result = CliRunner().invoke(cli, ["show", "A", "--root", str(fixtures_dir)])
    assert result.exit_code == 0
    assert "Tags:" in result.output
    assert "test" in result.output


def test_show_nested_requirement(fixtures_dir: Path) -> None:
    result = CliRunner().invoke(cli, ["show", "F", "--root", str(fixtures_dir)])
    assert result.exit_code == 0
    assert "Nested requirement" in result.output
    assert "Interface" in result.output


def test_show_omits_unset_optional_fields(fixtures_dir: Path) -> None:
    # A has no priority, status, stability, derived_from, or related_to
    result = CliRunner().invoke(cli, ["show", "A", "--root", str(fixtures_dir)])
    assert result.exit_code == 0
    assert "Priority:" not in result.output
    assert "Status:" not in result.output
    assert "Stability:" not in result.output
    assert "Derived From:" not in result.output
    assert "Related To:" not in result.output


def test_show_optional_fields_when_set(tmp_path: Path) -> None:
    _make_collection(tmp_path, [
        ("R1.md", """\
            ---
            id: R1
            title: Full requirement
            type: Functional
            verification: [Test]
            priority: High
            status: Active
            stability: Stable
            ---
            """),
    ])
    result = CliRunner().invoke(cli, ["show", "R1", "--root", str(tmp_path)])
    assert result.exit_code == 0
    assert "Priority:" in result.output
    assert "High" in result.output
    assert "Status:" in result.output
    assert "Active" in result.output
    assert "Stability:" in result.output
    assert "Stable" in result.output


# ---------------------------------------------------------------------------
# validate
# ---------------------------------------------------------------------------

def test_validate_exits_one_on_errors(tmp_path: Path) -> None:
    _make_collection(tmp_path, [
        ("R1.md", """\
            ---
            id: R1
            title: No verification
            type: Functional
            ---
            """),
    ])
    result = CliRunner().invoke(cli, ["validate", "--root", str(tmp_path)])
    assert result.exit_code == 1


def test_validate_reports_missing_required_field(tmp_path: Path) -> None:
    _make_collection(tmp_path, [
        ("R1.md", """\
            ---
            id: R1
            title: No verification
            type: Functional
            ---
            """),
    ])
    result = CliRunner().invoke(cli, ["validate", "--root", str(tmp_path)])
    assert "verification" in result.output.lower()


def test_validate_reports_duplicate_id(tmp_path: Path) -> None:
    _make_collection(tmp_path, [
        ("R1a.md", """\
            ---
            id: DUP
            title: First
            type: Functional
            verification: [Test]
            ---
            """),
        ("R1b.md", """\
            ---
            id: DUP
            title: Second
            type: Functional
            verification: [Test]
            ---
            """),
    ])
    result = CliRunner().invoke(cli, ["validate", "--root", str(tmp_path)])
    assert result.exit_code == 1
    assert "DUP" in result.output
    assert "Duplicate" in result.output


def test_validate_reports_broken_derived_from(tmp_path: Path) -> None:
    _make_collection(tmp_path, [
        ("R1.md", """\
            ---
            id: R1
            title: Orphaned child
            type: Functional
            verification: [Test]
            relationships:
              derived_from: [GHOST]
              related_to: []
            ---
            """),
    ])
    result = CliRunner().invoke(cli, ["validate", "--root", str(tmp_path)])
    assert result.exit_code == 1
    assert "GHOST" in result.output
    assert "derived_from" in result.output


def test_validate_reports_broken_related_to(tmp_path: Path) -> None:
    _make_collection(tmp_path, [
        ("R1.md", """\
            ---
            id: R1
            title: Bad reference
            type: Functional
            verification: [Test]
            relationships:
              derived_from: null
              related_to: [MISSING]
            ---
            """),
    ])
    result = CliRunner().invoke(cli, ["validate", "--root", str(tmp_path)])
    assert result.exit_code == 1
    assert "MISSING" in result.output
    assert "related_to" in result.output


def test_validate_reports_invalid_type(tmp_path: Path) -> None:
    _make_collection(tmp_path, [
        ("R1.md", """\
            ---
            id: R1
            title: Bad type
            type: Bogus
            verification: [Test]
            ---
            """),
    ])
    result = CliRunner().invoke(cli, ["validate", "--root", str(tmp_path)])
    assert result.exit_code == 1
    assert "Bogus" in result.output


def test_validate_reports_invalid_verification(tmp_path: Path) -> None:
    _make_collection(tmp_path, [
        ("R1.md", """\
            ---
            id: R1
            title: Bad verification
            type: Functional
            verification: [Guess]
            ---
            """),
    ])
    result = CliRunner().invoke(cli, ["validate", "--root", str(tmp_path)])
    assert result.exit_code == 1
    assert "Guess" in result.output


def test_validate_reports_unknown_key(tmp_path: Path) -> None:
    _make_collection(tmp_path, [
        ("R1.md", """\
            ---
            id: R1
            title: Unknown key
            type: Functional
            verification: [Test]
            alien_field: oops
            ---
            """),
    ])
    result = CliRunner().invoke(cli, ["validate", "--root", str(tmp_path)])
    assert result.exit_code == 1
    assert "alien_field" in result.output


def test_validate_all_violations_reported(tmp_path: Path) -> None:
    """Multiple violations in one collection all appear in a single run."""
    _make_collection(tmp_path, [
        ("R1.md", """\
            ---
            id: BAD
            title: Multiple problems
            type: NotAType
            verification: [Test]
            weird_key: stuff
            ---
            """),
    ])
    result = CliRunner().invoke(cli, ["validate", "--root", str(tmp_path)])
    assert result.exit_code == 1
    assert "NotAType" in result.output
    assert "weird_key" in result.output


# ---------------------------------------------------------------------------
# export
# ---------------------------------------------------------------------------

def test_export_default_output_filename(fixtures_dir: Path) -> None:
    """Omitting --output creates <exporter>.xlsx in the current directory."""
    runner = CliRunner()
    with runner.isolated_filesystem():
        result = runner.invoke(cli, ["export", "requirements", "--root", str(fixtures_dir)])
        assert result.exit_code == 0
        assert Path("requirements.xlsx").exists()


def test_export_requirements_sheet_names(tmp_path: Path, fixtures_dir: Path) -> None:
    output = tmp_path / "out.xlsx"
    CliRunner().invoke(cli, ["export", "requirements", "--root", str(fixtures_dir), "-o", str(output)])
    wb = openpyxl.load_workbook(output)
    assert "Requirements" in wb.sheetnames


def test_export_requirements_column_headers(tmp_path: Path, fixtures_dir: Path) -> None:
    output = tmp_path / "out.xlsx"
    CliRunner().invoke(cli, ["export", "requirements", "--root", str(fixtures_dir), "-o", str(output)])
    ws = openpyxl.load_workbook(output)["Requirements"]
    headers = [cell.value for cell in ws[1]]
    assert headers == [
        "Is Folder", "ID", "Parent", "Title", "Description", "Rationale",
        "Acceptance Criteria", "Type", "Verification Method",
        "Derived From", "Related To", "Tags", "Path",
    ]


def test_export_requirements_data_rows(tmp_path: Path, fixtures_dir: Path) -> None:
    output = tmp_path / "out.xlsx"
    CliRunner().invoke(cli, ["export", "requirements", "--root", str(fixtures_dir), "-o", str(output)])
    ws = openpyxl.load_workbook(output)["Requirements"]
    # Column 2 is ID (column 1 is Is Folder flag)
    ids = [ws.cell(row=r, column=2).value for r in range(2, ws.max_row + 1)]
    assert "A" in ids
    assert "F" in ids  # nested requirement must also be exported


def test_export_traceability_sheet_names(tmp_path: Path, fixtures_dir: Path) -> None:
    output = tmp_path / "out.xlsx"
    CliRunner().invoke(cli, ["export", "traceability", "--root", str(fixtures_dir), "-o", str(output)])
    wb = openpyxl.load_workbook(output)
    assert "Matrix" in wb.sheetnames
    assert "Orphans" in wb.sheetnames


def test_export_test_results_sheet_names(tmp_path: Path, fixtures_dir: Path) -> None:
    output = tmp_path / "out.xlsx"
    CliRunner().invoke(cli, ["export", "test-results", "--root", str(fixtures_dir), "-o", str(output)])
    wb = openpyxl.load_workbook(output)
    assert "Results" in wb.sheetnames
    assert "Coverage" in wb.sheetnames


def test_export_unknown_error_contains_name(fixtures_dir: Path) -> None:
    result = CliRunner().invoke(cli, ["export", "no-such-exporter", "--root", str(fixtures_dir)])
    assert result.exit_code == 1
    assert "no-such-exporter" in result.output


# ---------------------------------------------------------------------------
# Full workflow
# ---------------------------------------------------------------------------

def test_full_workflow(tmp_path: Path) -> None:
    """list → show → validate → export all succeed on a clean requirement set."""
    root = tmp_path / "reqs"
    root.mkdir()
    _make_collection(root, [
        ("SYS-001-performance.md", """\
            ---
            id: SYS-001
            title: System performance
            type: Performance
            verification: [Test, Analysis]
            priority: High
            status: Active
            ---

            ## Description

            The system shall respond within 200ms.

            ## Rationale

            Drives user experience quality.

            ## Acceptance Criteria

            Measured under peak load conditions.
            """),
        ("SYS-002-interface.md", """\
            ---
            id: SYS-002
            title: External interface
            type: Interface
            verification: [Inspection]
            relationships:
              derived_from: [SYS-001]
              related_to: []
            ---
            """),
    ])

    runner = CliRunner()

    # list: both requirements appear
    list_result = runner.invoke(cli, ["list", "--root", str(root)])
    assert list_result.exit_code == 0
    ids = [r.split()[0] for r in _data_lines(list_result.output)]
    assert "SYS-001" in ids
    assert "SYS-002" in ids

    # show: content of SYS-001 is present
    show_result = runner.invoke(cli, ["show", "SYS-001", "--root", str(root)])
    assert show_result.exit_code == 0
    assert "System performance" in show_result.output
    assert "200ms" in show_result.output
    assert "Priority:" in show_result.output
    assert "High" in show_result.output

    # validate: clean collection passes
    validate_result = runner.invoke(cli, ["validate", "--root", str(root)])
    assert validate_result.exit_code == 0
    assert "All requirements valid." in validate_result.output

    # export: workbook contains both IDs
    output = tmp_path / "report.xlsx"
    export_result = runner.invoke(
        cli, ["export", "requirements", "--root", str(root), "-o", str(output)]
    )
    assert export_result.exit_code == 0
    assert output.exists()
    ws = openpyxl.load_workbook(output)["Requirements"]
    # Column 2 is ID (column 1 is Is Folder flag)
    exported_ids = [ws.cell(row=r, column=2).value for r in range(2, ws.max_row + 1)]
    assert "SYS-001" in exported_ids
    assert "SYS-002" in exported_ids
